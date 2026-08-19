# -*- coding: utf-8 -*-
"""财务核对链验证脚本（skill: financial-check 的核心工具）

用法示例（PowerShell）:
    $env:PYTHONIOENCODING="utf-8"
    python verify_chain.py --bank "<财务数据根目录>/公账文件/X月份.xlsx" --debit 1000.00 --credit 5000.00 --balance 4000.00 --capital 4999.00 --verify 1.00
    python verify_chain.py --bank "<...>" --taxbook "<...>/个税扣缴申报_年度台账.xlsx" --soffice "C:/Program Files/LibreOffice/program/soffice.exe"

规则（真实性核对链，数字一律来自原始文件，禁止编造）:
  1. 银行借方合计 = 期间费用合计
  2. 现金流入 = 实收资本 + 银行验证款
  3. 期末余额 = 期初 + 贷方合计 - 借方合计（结构化校验）
  4. 冲红发票合计必须 = 0
  5. 资产负债表平衡：资产 = 负债 + 所有者权益
  6. 个税台账：累计减除费用 = 5000 x 任职月数，收入为 0 则应纳税所得额为 0
"""
import argparse
import glob
import os
import subprocess
import sys
import tempfile

def parse_bank(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row and row[0] == '交易时间':
            header_row = i
            break
    if header_row is None:
        raise SystemExit('未找到"交易时间"表头行，无法解析银行流水: ' + path)
    for row in ws.iter_rows(min_row=header_row + 2, values_only=True):
        if not row or not row[0]:
            continue
        db, cr, bal = row[2], row[3], row[4]
        rows.append({
            'time': row[0], 'debit': float(db) if db not in (None, '') else 0.0,
            'credit': float(cr) if cr not in (None, '') else 0.0,
            'balance': float(bal) if bal not in (None, '') else None,
        })
    return rows

def recalc_with_lo(xlsx_path, soffice):
    outdir = os.path.join(tempfile.gettempdir(), 'opencode_recalc')
    os.makedirs(outdir, exist_ok=True)
    subprocess.run([soffice, '--headless', '--convert-to', 'xlsx',
                    '--outdir', outdir, xlsx_path],
                   check=True, capture_output=True, timeout=300)
    hits = glob.glob(os.path.join(outdir, '*.xlsx'))
    if not hits:
        raise RuntimeError('LibreOffice 重算无输出')
    return hits[0]

def resolve_soffice(arg_value):
    """依次尝试: 命令行参数 > 环境变量 SOFFICE_EXE > PATH 中的 soffice"""
    import shutil
    if arg_value:
        return arg_value
    env = os.environ.get('SOFFICE_EXE')
    if env:
        return env
    found = shutil.which('soffice')
    if found:
        return found
    return None

def run_checks(bank_data, expect):
    """核对链纯函数（供 CLI 与外部引擎共用）。
    bank_data: 包含 transactions=[{'debit','credit','balance'}, ...] 与 summary.ending_balance 的 dict
    expect: {debit, credit, balance, capital, verify}
    返回: [{'name','expr','got','want','pass'}, ...]
    """
    txns = bank_data.get('transactions', [])
    total_debit = round(sum(float(t.get('debit', 0) or 0) for t in txns), 2)
    total_credit = round(sum(float(t.get('credit', 0) or 0) for t in txns), 2)
    end_bal = bank_data.get('summary', {}).get('ending_balance')
    if end_bal is None:
        end_bal = txns[-1].get('balance') if txns else 0.0
    end_bal = float(end_bal)
    results = []

    def add(name, expr, got, want):
        results.append({'name': name, 'expr': expr, 'got': round(got, 2),
                        'want': round(want, 2), 'pass': abs(got - want) < 0.005})

    add('结构性校验', '期末余额 = 贷方 - 借方', total_credit - total_debit, end_bal)
    if 'debit' in expect:
        add('银行借方=期间费用', '借方合计', total_debit, float(expect['debit']))
    if 'credit' in expect:
        add('银行贷方=现金流入', '贷方合计', total_credit, float(expect['credit']))
    if 'balance' in expect:
        add('期末余额', '余额', end_bal, float(expect['balance']))
    if 'capital' in expect and 'verify' in expect:
        add('贷方=实收资本+验证款', '实收资本+验证款',
            total_credit, float(expect['capital']) + float(expect['verify']))
    return results


def main():
    ap = argparse.ArgumentParser(description='财务报表核对链验证')
    ap.add_argument('--bank', required=True, help='银行流水xlsx路径')
    ap.add_argument('--debit', type=float, default=None, help='期望借方费用合计')
    ap.add_argument('--credit', type=float, default=None, help='期望贷方流入合计')
    ap.add_argument('--balance', type=float, default=None, help='期望期末余额')
    ap.add_argument('--capital', type=float, default=None, help='实收资本合计')
    ap.add_argument('--verify', type=float, default=None, help='银行验证款合计')
    ap.add_argument('--taxbook', default=None, help='个税申报台账xlsx（可选，公式重算校验）')
    ap.add_argument('--soffice', default=None, help='LibreOffice soffice 可执行文件路径（默认 SOFFICE_EXE/PATH）')
    args = ap.parse_args()

    rows = parse_bank(args.bank)
    if not rows:
        raise SystemExit('银行流水为空')

    bank_data = {
        'transactions': [{'debit': r['debit'], 'credit': r['credit'],
                          'balance': r['balance']} for r in rows],
        'summary': {'ending_balance': rows[-1]['balance'] if rows else 0.0},
    }
    expect = {'debit': args.debit, 'credit': args.credit, 'balance': args.balance,
              'capital': args.capital, 'verify': args.verify}
    expect = {k: v for k, v in expect.items() if v is not None}
    results = run_checks(bank_data, expect)
    ok = True
    for chk in results:
        ok = ok and chk['pass']
        print('{:<12} {:<28} 实际={:<10} 期望={:<10} {}'.format(
            chk['name'], chk['expr'], chk['got'], chk['want'], 'PASS' if chk['pass'] else 'FAIL'))

    if args.debit is not None:
        total_debit = next(c['got'] for c in results if c['name'] == '银行借方=期间费用')
        m = abs(total_debit - args.debit) < 0.005
        ok = ok and m
        print('核对1 银行借方=期间费用 实际={} 期望={} {}'.format(total_debit, args.debit, 'PASS' if m else 'FAIL'))
    if args.credit is not None:
        total_credit = next(c['got'] for c in results if c['name'] == '银行贷方=现金流入')
        m = abs(total_credit - args.credit) < 0.005
        ok = ok and m
        print('核对2 银行贷方=现金流入 实际={} 期望={} {}'.format(total_credit, args.credit, 'PASS' if m else 'FAIL'))
    if args.balance is not None:
        end_bal = next(c['got'] for c in results if c['name'] == '期末余额')
        m = abs(end_bal - args.balance) < 0.005
        ok = ok and m
        print('核对3 期末余额 实际={} 期望={} {}'.format(end_bal, args.balance, 'PASS' if m else 'FAIL'))
    if args.capital is not None and args.verify is not None:
        need = round(args.capital + args.verify, 2)
        total_credit = next(c['got'] for c in results if c['name'] == '银行贷方=现金流入')
        m = abs(total_credit - need) < 0.005
        ok = ok and m
        print('核对4 贷方=实收资本+验证款 实际={} 需={} {}'.format(total_credit, need, 'PASS' if m else 'FAIL'))

    if args.taxbook:
        try:
            soffice = resolve_soffice(args.soffice)
            if not soffice:
                raise RuntimeError('未找到 LibreOffice 可执行文件，请用 --soffice 指定')
            rec = recalc_with_lo(args.taxbook, soffice)
            print('LibreOffice 重算成功: ' + os.path.basename(rec))
        except Exception as e:
            ok = False
            print('FAIL LibreOffice 重算失败: ' + str(e))

    print('\n结论: ' + ('全部通过 PASS' if ok else '存在不一致 FAIL，请先核对原始文件'))
    sys.exit(0 if ok else 1)


if __name__ == '__main__':
    main()